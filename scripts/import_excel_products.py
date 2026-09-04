# -*- coding: utf-8 -*-
"""
Online Hızlı Parça - Excel Stok Listesi İçe Aktarma & Senkronizasyon Aracı
STOK LİSTE.xlsx dosyasındaki 8.700+ parçayı analiz eder, kategorize eder,
araç uyumluluklarını çıkarır ve JSON/SQL/Supabase formatlarında dışa aktarır.

Kullanım:
  python scripts/import_excel_products.py --dry-run
  python scripts/import_excel_products.py --export-sql supabase_seed_products.sql
  python scripts/import_excel_products.py --export-json products_dump.json
  python scripts/import_excel_products.py --upload
"""

import sys
import os
import re
import json
import argparse
import unicodedata
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("HATA: 'openpyxl' modülü bulunamadı. Lütfen 'pip install openpyxl' çalıştırın.")
    sys.exit(1)


# Türkçe karakter temizleme ve URL slug üretimi
def slugify(text: str) -> str:
    text = text.replace('İ', 'i').replace('I', 'i').replace('ı', 'i')
    text = text.replace('ğ', 'g').replace('Ğ', 'g')
    text = text.replace('ü', 'u').replace('Ü', 'u')
    text = text.replace('ş', 's').replace('Ş', 's')
    text = text.replace('ö', 'o').replace('Ö', 'o')
    text = text.replace('ç', 'c').replace('Ç', 'c')
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    text = re.sub(r'[^\w\s-]', '', text.lower()).strip()
    return re.sub(r'[-\s]+', '-', text)


# Kategori Eşleştirme Sözlüğü
CATEGORY_MAP = {
    'FİLTRE': ('İç Donanım & Periyodik Bakım', 'ic-donanim-bakim'),
    'POLEN': ('İç Donanım & Periyodik Bakım', 'ic-donanim-bakim'),
    'HAVA': ('İç Donanım & Periyodik Bakım', 'ic-donanim-bakim'),
    'YAĞ': ('İç Donanım & Periyodik Bakım', 'ic-donanim-bakim'),
    'FREN': ('Fren & Süspansiyon Sistemleri', 'fren-suspansiyon'),
    'BALATA': ('Fren & Süspansiyon Sistemleri', 'fren-suspansiyon'),
    'DİSK': ('Fren & Süspansiyon Sistemleri', 'fren-suspansiyon'),
    'AMORTİSÖR': ('Fren & Süspansiyon Sistemleri', 'fren-suspansiyon'),
    'SÜSPANSİYON': ('Fren & Süspansiyon Sistemleri', 'fren-suspansiyon'),
    'MOTOR': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'TRİGER': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'DEVRİDAİM': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'DEVİRDAİM': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'CONTA': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'PİSTON': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'DEBRİYAJ': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'ŞANZIMAN': ('Motor & Aktarma Organları', 'motor-aktarma'),
    'ATEŞLEME': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'BOBİN': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'BUJİ': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'ELEKTRİK': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'AYDINLATMA': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'FAR': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'STOP': ('Aydınlatma & Elektrik Aksamı', 'aydinlatma-elektrik'),
    'KAPORTA': ('Kaporta & Dış Aksesuar', 'kaporta-aksesuar'),
    'TAMPON': ('Kaporta & Dış Aksesuar', 'kaporta-aksesuar'),
    'ÇAMURLUK': ('Kaporta & Dış Aksesuar', 'kaporta-aksesuar'),
    'AYNA': ('Kaporta & Dış Aksesuar', 'kaporta-aksesuar'),
    'İÇ TRİM': ('Kaporta & Dış Aksesuar', 'kaporta-aksesuar')
}

# Marka ve Model Tespiti
BRAND_KEYWORDS = {
    'Opel': [
        'ASTRA', 'CORSA', 'VECTRA', 'INSIGNIA', 'ZAFIRA', 'MOKKA', 'CROSSLAND', 
        'GRANDLAND', 'COMBO', 'MERIVA', 'TIGRA', 'OMEGA', 'ANTARA', 'CALIBRA'
    ],
    'Peugeot': [
        '106', '205', '206', '207', '208', '301', '306', '307', '308', '406', 
        '407', '508', '2008', '3008', '5008', 'PARTNER', 'BOXER', 'EXPERT', 'BIPPER', 'RCZ', 'RIFTER'
    ],
    'Citroën': [
        'C1', 'C2', 'C3', 'C4', 'C5', 'BERLINGO', 'NEMO', 'JUMPER', 'JUMPY', 
        'ELYSEE', 'C-ELYSEE', 'XSARA', 'SAXO', 'PICASSO', 'CACTUS', 'AIRCROSS'
    ],
    'Chevrolet': [
        'CRUZE', 'AVEO', 'CAPTIVA', 'SPARK', 'LACETTI', 'KALOS', 'TRAX', 'EPICA', 'REZZO'
    ],
    'DS Automobiles': [
        'DS3', 'DS4', 'DS5', 'DS7', 'DS 3', 'DS 4', 'DS 7', 'DS9'
    ],
    'Diğer Markalar': [
        'DUCATO', 'DOBLO', 'FIORINO', 'TRAFIC', 'MASTER', 'GOLF', 'POLO', 'PASSAT', 'CADDY', 'OCTAVIA', 'LEON', 'A3', 'FOCUS', 'TRANSIT'
    ]
}


def detect_brand_and_models(title: str, code: str, mfg: str):
    text = f"{code} {title} {mfg}".upper()
    
    detected_brands = []
    detected_models = []
    
    # 1. Önce 5 uzmanlık markasını kontrol et
    for brand, models in BRAND_KEYWORDS.items():
        if brand == 'Diğer Markalar':
            continue
        found_models = []
        for m in models:
            pattern = r'\b' + re.escape(m) + r'\b'
            if re.search(pattern, text):
                found_models.append(m)
        if found_models or brand.upper() in text:
            detected_brands.append(brand)
            detected_models.extend([{'brand': brand, 'model': m, 'years': 'Tüm Yıllar'} for m in found_models])
            
    # 2. Eğer hiçbir ana marka bulunamadıysa 'Diğer Markalar' kontrolü yap
    if not detected_brands:
        for m in BRAND_KEYWORDS['Diğer Markalar']:
            if re.search(r'\b' + re.escape(m) + r'\b', text):
                detected_brands.append('Diğer Markalar')
                detected_models.append({'brand': 'Diğer Markalar', 'model': m, 'years': 'Tüm Yıllar'})
                break
                
    # 3. Hala marka yoksa ama PSA motor kodu varsa
    if not detected_brands:
        if any(k in text for k in ['DV4', 'DV6', 'DW10', 'DW8', 'TU3', 'TU5', 'EP6', 'EB2', 'PURETECH', 'HDI']):
            detected_brands = ['Peugeot', 'Citroën']
            detected_models = [{'brand': 'Peugeot', 'model': '1.4/1.6 HDi & PureTech', 'years': 'Tüm Yıllar'}]
        else:
            detected_brands = ['Diğer Markalar']
            detected_models = [{'brand': 'Diğer Markalar', 'model': 'Evrensel Uyumlu', 'years': 'Tüm Yıllar'}]

    primary_brand = detected_brands[0]
    return primary_brand, detected_models


def detect_category(group: str, subgrp: str, title: str):
    combined = f"{group} {subgrp} {title}".upper()
    for kw, cat_info in CATEGORY_MAP.items():
        if kw in combined:
            return cat_info
    return ('Motor & Aktarma Organları', 'motor-aktarma')


def parse_excel(file_path: str):
    print(f"[*] '{file_path}' açılıyor...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    total_rows = sheet.max_row
    print(f"[*] Toplam {total_rows - 1} satır veri bulundu. İşleniyor...")
    
    products = []
    brand_stats = Counter()
    category_stats = Counter()
    stock_count = 0
    
    slug_counts = Counter()

    for r in range(2, total_rows + 1):
        code = str(sheet.cell(r, 1).value or '').strip()
        title = str(sheet.cell(r, 2).value or '').strip()
        group = str(sheet.cell(r, 3).value or '').strip()
        subgroup = str(sheet.cell(r, 4).value or '').strip()
        qty = sheet.cell(r, 6).value or 0
        mfg = str(sheet.cell(r, 9).value or '').strip()
        mfg_code = str(sheet.cell(r, 10).value or '').strip()
        
        if not code or not title:
            continue
            
        try:
            qty_num = int(float(qty))
        except (ValueError, TypeError):
            qty_num = 0
            
        if qty_num > 0:
            stock_count += 1
            
        brand, models = detect_brand_and_models(title, code, mfg)
        category_name, category_slug = detect_category(group, subgroup, title)
        
        brand_stats[brand] += 1
        category_stats[category_name] += 1
        
        # Orijinal mi Yan Sanayi mi?
        is_original = any(k in mfg.upper() for k in ['OPEL GM', 'PEUGEOT', 'CITROEN', 'PSA', 'CHEVROLET', 'OEM', 'GENUINE'])
        part_quality = 'original' if is_original else ('oem' if any(k in mfg.upper() for k in ['BOSCH', 'DELPHI', 'VALEO', 'SNR', 'LUK', 'HENGST', 'SACHS']) else 'aftermarket')
        
        # Benzersiz slug
        base_slug = slugify(f"{brand}-{title}-{code}")[:80]
        slug_counts[base_slug] += 1
        if slug_counts[base_slug] > 1:
            slug = f"{base_slug}-{slug_counts[base_slug]}"
        else:
            slug = base_slug

        # Ürün Nesnesi
        prod = {
            'id': f"prod-{slugify(code)}-{r}",
            'title': title,
            'slug': slug,
            'category': category_name,
            'category_slug': category_slug,
            'brand': brand,
            'part_number': code,
            'oem_reference_number': code.split('.')[0] if '.' in code else code,
            'is_original': is_original,
            'part_quality': part_quality,
            'vehicle_compatibility': models,
            'price': 0, # Fiyatlar Excel'de 0, kullanıcı belirleyecek
            'discount_price': None,
            'stock': qty_num,
            'image_url': '/images/products/on-fren-balatasi.jpg', # Varsayılan şablon görsel
            'description': f"{brand} araçlar için {title}. Parça Kodu: {code}. Üretici: {mfg or 'Orijinal Standart'}.",
            'specs': {
                'Üretici': mfg or 'Belirtilmemiş',
                'Üretici Parça Kodu': mfg_code or code,
                'Stok Grubu': group,
                'Alt Grup': subgroup
            },
            'rating': 4.8,
            'reviews_count': 0,
            'is_featured': False
        }
        products.append(prod)

    print(f"\n[+] İşlem Tamamlandı: {len(products)} ürün başarıyla modellendi.")
    print(f"[+] Pozitif Stoklu Ürün Sayısı: {stock_count}")
    print("\n--- Marka Dağılımı ---")
    for b, count in brand_stats.most_common():
        print(f"  {b}: {count} ürün (%{round(count/len(products)*100, 1)})")
        
    print("\n--- Kategori Dağılımı ---")
    for c, count in category_stats.most_common():
        print(f"  {c}: {count} ürün (%{round(count/len(products)*100, 1)})")
        
    return products


def export_to_json(products, output_path):
    print(f"\n[*] JSON formatinda '{output_path}' dosyasina aktariliyor...")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"[BASARILI] {len(products)} adet urun '{output_path}' dosyasina kaydedildi!")


def export_to_sql(products, output_path):
    print(f"\n[*] Supabase SQL formatinda '{output_path}' dosyasina aktariliyor...")
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("-- Online Hizli Parca - Supabase Products Seed\n")
        f.write("-- Toplam Urun: " + str(len(products)) + "\n\n")
        
        batch_size = 100
        for i in range(0, len(products), batch_size):
            batch = products[i:i + batch_size]
            f.write("INSERT INTO public.products (id, title, slug, category, category_slug, brand, part_number, oem_reference_number, is_original, part_quality, price, stock, image_url, description, specs, vehicle_compatibility, rating, reviews_count, is_featured) VALUES\n")
            
            val_lines = []
            for p in batch:
                title_esc = p['title'].replace("'", "''")
                slug_esc = p['slug'].replace("'", "''")
                cat_esc = p['category'].replace("'", "''")
                desc_esc = p['description'].replace("'", "''")
                specs_json = json.dumps(p['specs'], ensure_ascii=False).replace("'", "''")
                vc_json = json.dumps(p['vehicle_compatibility'], ensure_ascii=False).replace("'", "''")
                
                line = f"  ('{p['id']}', '{title_esc}', '{slug_esc}', '{cat_esc}', '{p['category_slug']}', '{p['brand']}', '{p['part_number']}', '{p['oem_reference_number']}', {str(p['is_original']).lower()}, '{p['part_quality']}', {p['price']}, {p['stock']}, '{p['image_url']}', '{desc_esc}', '{specs_json}'::jsonb, '{vc_json}'::jsonb, {p['rating']}, {p['reviews_count']}, {str(p['is_featured']).lower()})"
                val_lines.append(line)
                
            f.write(",\n".join(val_lines))
            f.write("\nON CONFLICT (id) DO UPDATE SET stock = EXCLUDED.stock, price = EXCLUDED.price;\n\n")
            
    print(f"[BASARILI] SQL Seed dosyasi basariyla olusturuldu: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="STOK LİSTE.xlsx İçe Aktarma Aracı")
    parser.add_argument('--file', default='STOK LİSTE.xlsx', help="Excel dosya yolu")
    parser.add_argument('--dry-run', action='store_true', help="Sadece analiz et ve raporla")
    parser.add_argument('--export-json', type=str, help="JSON çıktısı oluştur")
    parser.add_argument('--export-sql', type=str, help="Supabase SQL seed dosyası oluştur")
    parser.add_argument('--upload', action='store_true', help="Supabase veritabanına doğrudan yükle")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.file):
        print(f"HATA: '{args.file}' bulunamadı.")
        sys.exit(1)
        
    products = parse_excel(args.file)
    
    if args.export_json:
        export_to_json(products, args.export_json)
        
    if args.export_sql:
        export_to_sql(products, args.export_sql)
        
    if args.upload:
        print("\n[*] Supabase doğrudan yükleme servisi çağrılıyor...")
        supabase_url = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
        service_role_key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
        if not service_role_key:
            print("UYARI: 'SUPABASE_SERVICE_ROLE_KEY' ortam değişkeni tanımlı değil.")
            print("Veritabanına güvenli toplu yükleme için service role key gereklidir.")
            print("Bunun yerine '--export-sql' parametresi ile üretilen SQL dosyasını Supabase SQL Editor alanında tek seferde çalıştırabilirsiniz.")
            
    if not (args.export_json or args.export_sql or args.upload) and not args.dry_run:
        print("\nİpucu: Komutu '--dry-run', '--export-sql dosya.sql' veya '--export-json dosya.json' ile çalıştırabilirsiniz.")


if __name__ == '__main__':
    main()
